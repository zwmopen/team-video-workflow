from __future__ import annotations

from dataclasses import dataclass, asdict
from pathlib import Path
import csv
import html
import json
import re

from .ffmpeg_utils import find_ffmpeg, find_ffprobe, probe_video, run_command
from .path_utils import ensure_unique_path, sanitize_name, seconds_to_timestamp
from .script_matcher import KEYWORD_TARGETS, ScriptBeat, load_clip_records, make_beat, split_script


@dataclass(slots=True)
class MatchBoardOptions:
    library_root: Path
    title: str
    output: Path | None = None
    script_file: Path | None = None
    script_text: str = ""
    audio_file: Path | None = None
    reference_video: Path | None = None
    feedback_file: Path | None = None
    max_candidates: int = 5


@dataclass(slots=True)
class MatchCandidate:
    rank: int
    clip_path: str
    thumbnail: str
    category: str
    keyword: str
    source_video: str
    score: int
    reason: str
    direct: bool


@dataclass(slots=True)
class MatchBoardRow:
    index: int
    start: float | None
    end: float | None
    script_text: str
    visual_need: str
    target_category: str
    target_keyword: str
    candidates: list[MatchCandidate]
    status: str
    note: str = ""


CONCRETE_SUPPORT_CATEGORIES = {
    "01_环境空镜",
    "06_团队互动",
    "08_人物反应",
    "09_细节特写",
}


def build_shot_match_board(options: MatchBoardOptions) -> dict[str, object]:
    library_root = options.library_root.expanduser().resolve()
    if not library_root.exists():
        raise FileNotFoundError(f"Library root does not exist: {library_root}")
    output = options.output or (library_root.parent / f"{sanitize_name(options.title)}_智能镜头匹配审片板")
    output = output.expanduser().resolve()
    output.mkdir(parents=True, exist_ok=True)
    thumbs_dir = output / "候选截图"
    thumbs_dir.mkdir(parents=True, exist_ok=True)

    script_text = resolve_board_script_text(options)
    beats = parse_timecoded_beats(script_text)
    if not beats:
        beats = untimed_rows_from_script(script_text)
    records = load_clip_records(library_root)
    feedback_by_index = load_board_feedback(options.feedback_file)
    rows: list[MatchBoardRow] = []
    used_first_choices: set[str] = set()
    last_source = ""
    ffmpeg = find_ffmpeg()
    ffprobe = find_ffprobe()

    for beat in beats:
        beat = apply_feedback_replacement_to_beat(beat, feedback_by_index.get(beat.index, {}))
        candidates = select_board_candidates(
            records=records,
            beat=beat,
            output_dir=output,
            thumbs_dir=thumbs_dir,
            max_candidates=options.max_candidates,
            used_first_choices=used_first_choices,
            last_source=last_source,
            ffmpeg=ffmpeg,
            ffprobe=ffprobe,
        )
        if candidates:
            used_first_choices.add(candidates[0].clip_path)
            last_source = candidates[0].source_video
        status = "待审" if candidates and candidates[0].direct else "缺直接素材"
        rows.append(
            MatchBoardRow(
                index=beat.index,
                start=getattr(beat, "start", None),
                end=getattr(beat, "end", None),
                script_text=beat.text,
                visual_need=visual_need_for_beat(beat),
                target_category=beat.target_category,
                target_keyword=beat.target_subcategory,
                candidates=candidates,
                status=status,
            )
        )

    audio_path = ""
    transcript_path = ""
    if options.audio_file:
        audio = options.audio_file.expanduser().resolve()
        if audio.exists():
            audio_path = str(audio)
            transcript = audio.with_suffix(".txt")
            if transcript.exists():
                transcript_path = str(transcript)
    elif options.reference_video:
        audio_path, transcript_path = export_reference_audio_and_text(options.reference_video, script_text, output, ffmpeg)

    csv_path = output / "智能镜头匹配候选表.csv"
    xlsx_path = output / "智能镜头匹配审片表.xlsx"
    json_path = output / "智能镜头匹配候选表.json"
    html_path = output / "智能镜头匹配审片板.html"
    feedback_csv_path = output / "智能镜头匹配审片反馈表.csv"
    feedback_json_path = output / "智能镜头匹配审片反馈表.json"
    write_match_board_csv(csv_path, rows, options.max_candidates, output)
    write_match_board_xlsx(xlsx_path, rows, options.max_candidates, output)
    write_feedback_template(feedback_csv_path, feedback_json_path, rows)
    json_path.write_text(
        json.dumps(
            {
                "title": options.title,
                "library_root": str(library_root),
                "audio_file": str(options.audio_file or ""),
                "reference_video": str(options.reference_video or ""),
                "input_feedback_file": str(options.feedback_file or ""),
                "audio_path": audio_path,
                "transcript_path": transcript_path,
                "feedback_csv": str(feedback_csv_path),
                "feedback_json": str(feedback_json_path),
                "review_xlsx": str(xlsx_path),
                "rows": [row_to_json(row, output) for row in rows],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    write_match_board_html(html_path, options.title, rows, output, audio_path)
    return {
        "library_root": str(library_root),
        "output": str(output),
        "html": str(html_path),
        "csv": str(csv_path),
        "xlsx": str(xlsx_path),
        "json": str(json_path),
        "feedback_csv": str(feedback_csv_path),
        "feedback_json": str(feedback_json_path),
        "audio_path": audio_path,
        "beats": len(rows),
        "ready": sum(1 for row in rows if row.status == "待审"),
        "missing_direct": sum(1 for row in rows if row.status != "待审"),
    }


def resolve_board_script_text(options: MatchBoardOptions) -> str:
    if options.script_file:
        return clean_script_text(options.script_file.expanduser().read_text(encoding="utf-8", errors="replace"))
    if options.script_text.strip():
        return clean_script_text(options.script_text)
    if options.audio_file:
        sidecar = options.audio_file.expanduser().with_suffix(".txt")
        if sidecar.exists():
            return clean_script_text(sidecar.read_text(encoding="utf-8", errors="replace"))
    if options.reference_video:
        sidecar = options.reference_video.expanduser().with_suffix(".txt")
        if sidecar.exists():
            return clean_script_text(sidecar.read_text(encoding="utf-8", errors="replace"))
    raise ValueError("Provide --script-file, --script-text, --audio-file with same-stem .txt, or a reference video with same-stem .txt")


def clean_script_text(text: str) -> str:
    lines = []
    for raw_line in text.splitlines():
        raw_line = raw_line.lstrip("\ufeff")
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        lines.append(raw_line)
    return "\n".join(lines).strip()


def load_board_feedback(path: Path | None) -> dict[int, dict[str, str]]:
    if not path:
        return {}
    feedback_path = path.expanduser().resolve()
    if not feedback_path.exists():
        raise FileNotFoundError(f"Feedback file does not exist: {feedback_path}")
    if feedback_path.suffix.lower() == ".json":
        raw = json.loads(feedback_path.read_text(encoding="utf-8-sig"))
        rows = raw if isinstance(raw, list) else raw.get("rows", [])
        if not isinstance(rows, list):
            return {}
        return index_feedback_rows(rows)
    with feedback_path.open("r", encoding="utf-8-sig", newline="") as handle:
        return index_feedback_rows(list(csv.DictReader(handle)))


def index_feedback_rows(rows: list[object]) -> dict[int, dict[str, str]]:
    indexed: dict[int, dict[str, str]] = {}
    for raw_row in rows:
        if not isinstance(raw_row, dict):
            continue
        row = {str(key).strip(): "" if value is None else str(value).strip() for key, value in raw_row.items()}
        index_value = feedback_value(row, ["镜号", "镜头号", "序号", "index", "beat_index", "row"])
        if not index_value:
            continue
        match = re.search(r"\d+", index_value)
        if not match:
            continue
        indexed[int(match.group(0))] = row
    return indexed


def feedback_value(row: dict[str, str], keys: list[str]) -> str:
    normalized = {normalize_feedback_key(key): value for key, value in row.items()}
    for key in keys:
        value = normalized.get(normalize_feedback_key(key), "")
        if value:
            return value.strip()
    return ""


def normalize_feedback_key(value: str) -> str:
    return re.sub(r"\s+", "", value.strip().lower())


def feedback_replacement_keyword(row: dict[str, str]) -> str:
    return feedback_value(
        row,
        [
            "替换关键词",
            "重配关键词",
            "新关键词",
            "画面关键词",
            "replacement_keyword",
            "replace_keyword",
            "keyword",
        ],
    )


def apply_feedback_replacement_to_beat(beat: "TimedBeat", feedback_row: dict[str, str]) -> "TimedBeat":
    replacement = feedback_replacement_keyword(feedback_row)
    if not replacement:
        return beat
    inferred = make_beat(beat.index, f"{replacement} {beat.text}")
    target_category = inferred.target_category
    target_keyword = inferred.target_subcategory or replacement
    if target_category.startswith("90_") and not inferred.target_subcategory:
        target_category = beat.target_category
    return TimedBeat(
        index=beat.index,
        text=beat.text,
        target_category=target_category,
        target_subcategory=target_keyword,
        suggested_duration=beat.suggested_duration,
        start=beat.start,
        end=beat.end,
    )


@dataclass(slots=True)
class TimedBeat(ScriptBeat):
    start: float | None = None
    end: float | None = None


def parse_timecoded_beats(text: str) -> list[TimedBeat]:
    rows: list[TimedBeat] = []
    pending_start: float | None = None
    pending_end: float | None = None
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line.isdigit():
            continue
        short_arrow_match = re.match(
            r"(?P<start>\d{1,2}:\d{2}(?::\d{2})?(?:[,.]\d{1,3})?)\s*-->\s*(?P<end>\d{1,2}:\d{2}(?::\d{2})?(?:[,.]\d{1,3})?)\s*(?P<text>.*)$",
            line,
        )
        if short_arrow_match:
            text_value = short_arrow_match.group("text").strip(" \t-—：:")
            if text_value:
                rows.append(
                    make_timed_beat(
                        len(rows) + 1,
                        text_value,
                        parse_timestamp(short_arrow_match.group("start")),
                        parse_timestamp(short_arrow_match.group("end")),
                    )
                )
            else:
                pending_start = parse_timestamp(short_arrow_match.group("start"))
                pending_end = parse_timestamp(short_arrow_match.group("end"))
            continue
        srt_match = re.match(r"(?P<start>\d{1,2}:\d{2}:\d{2}[,.]\d{1,3})\s*-->\s*(?P<end>\d{1,2}:\d{2}:\d{2}[,.]\d{1,3})", line)
        if srt_match:
            pending_start = parse_timestamp(srt_match.group("start"))
            pending_end = parse_timestamp(srt_match.group("end"))
            rest = line[srt_match.end() :].strip(" \t-—:：")
            if rest:
                rows.append(make_timed_beat(len(rows) + 1, rest, pending_start, pending_end))
                pending_start = None
                pending_end = None
            continue
        inline_match = re.match(
            r"[\[\(【]?(?P<start>\d{1,2}:\d{2}(?::\d{2})?(?:[,.]\d{1,3})?)\s*(?:-->|-|—|~|～|至)\s*(?P<end>\d{1,2}:\d{2}(?::\d{2})?(?:[,.]\d{1,3})?)[\]\)】]?\s*(?P<text>.+)$",
            line,
        )
        if inline_match:
            rows.append(
                make_timed_beat(
                    len(rows) + 1,
                    inline_match.group("text").strip(),
                    parse_timestamp(inline_match.group("start")),
                    parse_timestamp(inline_match.group("end")),
                )
            )
            pending_start = None
            pending_end = None
            continue
        bracket_match = re.match(r"[\[\(【]?(?P<start>\d{1,2}:\d{2}(?::\d{2})?(?:[,.]\d{1,3})?)[\]\)】]\s*(?P<text>.+)$", line)
        if bracket_match:
            start = parse_timestamp(bracket_match.group("start"))
            rows.append(make_timed_beat(len(rows) + 1, bracket_match.group("text").strip(), start, None))
            pending_start = None
            pending_end = None
            continue
        if pending_start is not None:
            rows.append(make_timed_beat(len(rows) + 1, line, pending_start, pending_end))
            pending_start = None
            pending_end = None
    return [row for row in rows if row.text]


def untimed_rows_from_script(text: str) -> list[TimedBeat]:
    return [make_timed_beat(beat.index, beat.text, None, None) for beat in split_script(text)]


def make_timed_beat(index: int, text: str, start: float | None, end: float | None) -> TimedBeat:
    beat = make_beat(index, text)
    return TimedBeat(
        index=index,
        text=beat.text,
        target_category=beat.target_category,
        target_subcategory=beat.target_subcategory,
        suggested_duration=beat.suggested_duration,
        start=start,
        end=end,
    )


def parse_timestamp(value: str) -> float:
    cleaned = value.replace(",", ".")
    parts = cleaned.split(":")
    if len(parts) == 2:
        minute, second = parts
        return int(minute) * 60 + float(second)
    if len(parts) == 3:
        hour, minute, second = parts
        return int(hour) * 3600 + int(minute) * 60 + float(second)
    return 0.0


def visual_need_for_beat(beat: ScriptBeat) -> str:
    keyword = beat.target_subcategory or beat.target_category.split("_", 1)[-1]
    if beat.target_category.startswith("90_"):
        keyword = infer_keyword_from_text(beat.text) or "待人工判断"
    return keyword


def infer_keyword_from_text(text: str) -> str:
    for pattern, _, subcategory in KEYWORD_TARGETS:
        if re.search(pattern, text, flags=re.IGNORECASE):
            return subcategory
    return ""


def select_board_candidates(
    records: list[dict[str, object]],
    beat: TimedBeat,
    output_dir: Path,
    thumbs_dir: Path,
    max_candidates: int,
    used_first_choices: set[str],
    last_source: str,
    ffmpeg: Path,
    ffprobe: Path | None,
) -> list[MatchCandidate]:
    scored = []
    direct_found = False
    for record in records:
        path = Path(str(record.get("output_path", "")))
        if not path.exists():
            continue
        score, reason, direct = score_record_for_beat(record, beat, used_first_choices, last_source)
        if score <= 0:
            continue
        direct_found = direct_found or direct
        scored.append((score, direct, record, reason))
    if direct_found:
        scored = [item for item in scored if item[1] or str(item[2].get("primary_category", "")) in CONCRETE_SUPPORT_CATEGORIES]
    scored.sort(key=lambda item: (item[1], item[0]), reverse=True)
    selected: list[MatchCandidate] = []
    seen_paths: set[str] = set()
    for score, direct, record, reason in scored:
        path = str(record.get("output_path", ""))
        if path in seen_paths:
            continue
        seen_paths.add(path)
        rank = len(selected) + 1
        thumb_path = thumbs_dir / f"{beat.index:03d}_{rank:02d}_{sanitize_name(Path(path).stem)}.jpg"
        make_candidate_thumbnail(Path(path), thumb_path, ffmpeg, ffprobe)
        selected.append(
            MatchCandidate(
                rank=rank,
                clip_path=path,
                thumbnail=relative_or_absolute(thumb_path, output_dir),
                category=infer_category_from_path(path) or str(record.get("primary_category", "")),
                keyword=infer_keyword_from_path(path) or str(record.get("category_top1", "")),
                source_video=str(record.get("source_video_name", "")) or str(record.get("source_video_path", "")),
                score=score,
                reason=reason,
                direct=direct,
            )
        )
        if len(selected) >= max_candidates:
            break
    return selected


def score_record_for_beat(
    record: dict[str, object],
    beat: TimedBeat,
    used_first_choices: set[str],
    last_source: str,
) -> tuple[int, str, bool]:
    path = str(record.get("output_path", ""))
    primary = infer_category_from_path(path) or str(record.get("primary_category", ""))
    keyword = infer_keyword_from_path(path) or str(record.get("category_top1", ""))
    physical_evidence = " ".join([path, primary, keyword])
    semantic = " ".join([physical_evidence, str(record.get("semantic_tags", ""))])
    source = str(record.get("source_video_name", "")) or str(record.get("source_video_path", ""))
    quality = str(record.get("quality_level", "B"))
    target_keyword = beat.target_subcategory.strip()
    target_words = [target_keyword, beat.target_category.split("_", 1)[-1], infer_keyword_from_text(beat.text)]
    score = 0
    reasons: list[str] = []
    direct = False
    if primary == beat.target_category:
        score += 35
        reasons.append("一级分类命中")
    if target_keyword and target_keyword in physical_evidence:
        score += 55
        direct = True
        reasons.append(f"具体关键词命中：{target_keyword}")
    elif target_keyword and target_keyword in str(record.get("semantic_tags", "")):
        score += 12
        reasons.append("历史记录标签弱命中")
    elif any(word and word in semantic for word in target_words):
        score += 18
        reasons.append("文本关键词弱命中")
    if target_keyword and primary == beat.target_category and not direct:
        return 0, "", False
    if target_keyword and primary != beat.target_category and not direct:
        if primary not in CONCRETE_SUPPORT_CATEGORIES:
            return 0, "", False
        score += 4
        reasons.append("辅助镜头候选")
    if beat.target_category.startswith("90_") and score <= 0:
        return 0, "", False
    if path in used_first_choices:
        score -= 18
        reasons.append("已作为前面首选，降权")
    if source and source == last_source:
        score -= 8
        reasons.append("与上一句同源，降权")
    if quality == "S":
        score += 8
        reasons.append("质量 S")
    elif quality == "A":
        score += 5
        reasons.append("质量 A")
    elif quality == "B":
        score += 2
    return score, "；".join(reasons) or "路径/标签相关", direct or (primary == beat.target_category and not target_keyword)


def infer_keyword_from_path(path: str) -> str:
    parts = Path(path).parts
    for index, part in enumerate(parts):
        if re.match(r"^\d{2}_", part) and index + 1 < len(parts):
            return parts[index + 1]
    return Path(path).parent.name


def infer_category_from_path(path: str) -> str:
    for part in Path(path).parts:
        if re.match(r"^\d{2}_", part):
            return part
    return ""


def make_candidate_thumbnail(video: Path, thumb_path: Path, ffmpeg: Path, ffprobe: Path | None) -> None:
    if thumb_path.exists() and thumb_path.stat().st_size > 0:
        return
    thumb_path.parent.mkdir(parents=True, exist_ok=True)
    duration = 0.0
    try:
        duration = float(probe_video(video, ffmpeg, ffprobe).get("duration") or 0.0)
    except Exception:
        duration = 0.0
    seek = max(0.05, min(duration * 0.5, 1.0)) if duration else 0.15
    result = run_command(
        [
            ffmpeg,
            "-hide_banner",
            "-loglevel",
            "error",
            "-y",
            "-ss",
            f"{seek:.3f}",
            "-i",
            video,
            "-frames:v",
            "1",
            "-vf",
            "scale=240:-1",
            "-q:v",
            "3",
            thumb_path,
        ],
        timeout=45,
    )
    if result.returncode != 0:
        thumb_path.unlink(missing_ok=True)


def export_reference_audio_and_text(reference_video: Path, script_text: str, output: Path, ffmpeg: Path) -> tuple[str, str]:
    source = reference_video.expanduser().resolve()
    if not source.exists():
        return "", ""
    audio_dir = output / "音频提取"
    audio_dir.mkdir(parents=True, exist_ok=True)
    audio_path = ensure_unique_path(audio_dir / f"{sanitize_name(source.stem)}.m4a")
    result = run_command(
        [ffmpeg, "-y", "-i", source, "-vn", "-c:a", "aac", "-b:a", "192k", audio_path],
        timeout=300,
    )
    if result.returncode != 0:
        audio_path.unlink(missing_ok=True)
        audio_path = Path("")
    transcript_path = audio_dir / f"{sanitize_name(source.stem)}.txt"
    transcript_path.write_text(script_text, encoding="utf-8")
    return (str(audio_path) if str(audio_path) != "." else ""), str(transcript_path)


def write_match_board_csv(path: Path, rows: list[MatchBoardRow], max_candidates: int, output_dir: Path) -> None:
    fields = ["镜号", "开始", "结束", "字幕台词", "AI画面需求", "目标分类", "目标关键词", "状态"]
    for index in range(1, max_candidates + 1):
        fields.extend([f"素材{index}名称", f"素材{index}截图", f"素材{index}理由", f"素材{index}路径"])
    fields.append("备注")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            data = {
                "镜号": f"{row.index:03d}",
                "开始": seconds_to_timestamp(row.start) if row.start is not None else "",
                "结束": seconds_to_timestamp(row.end) if row.end is not None else "",
                "字幕台词": row.script_text,
                "AI画面需求": row.visual_need,
                "目标分类": row.target_category,
                "目标关键词": row.target_keyword,
                "状态": row.status,
                "备注": row.note,
            }
            for candidate in row.candidates:
                data[f"素材{candidate.rank}名称"] = Path(candidate.clip_path).name
                data[f"素材{candidate.rank}截图"] = candidate.thumbnail
                data[f"素材{candidate.rank}理由"] = candidate.reason
                data[f"素材{candidate.rank}路径"] = candidate.clip_path
            writer.writerow(data)


def write_match_board_xlsx(path: Path, rows: list[MatchBoardRow], max_candidates: int, output_dir: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XlsxImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "智能镜头匹配"
    headers = ["镜号", "时间", "字幕台词", "AI画面需求", "目标分类", "目标关键词", "状态"]
    for index in range(1, max_candidates + 1):
        headers.extend([f"素材{index}截图", f"素材{index}名称/理由"])
    headers.extend(["确认结果", "选中素材序号", "替换关键词", "备注"])
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="DCEBFF")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="1C2938")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start=2):
        time_text = ""
        if row.start is not None or row.end is not None:
            time_text = f"{seconds_to_timestamp(row.start or 0)} - {seconds_to_timestamp(row.end or 0)}"
        values = [
            f"{row.index:03d}",
            time_text,
            row.script_text,
            row.visual_need,
            row.target_category,
            row.target_keyword,
            row.status,
        ]
        candidate_text_start = len(values) + 2
        for candidate in row.candidates[:max_candidates]:
            values.extend(["", f"{Path(candidate.clip_path).name}\n{candidate.category} / {candidate.keyword}\n{candidate.reason}"])
        for _ in range(max_candidates - len(row.candidates[:max_candidates])):
            values.extend(["", ""])
        values.extend(["", "", "", ""])
        sheet.append(values)
        sheet.row_dimensions[row_index].height = 116

        for candidate in row.candidates[:max_candidates]:
            image_column = 8 + (candidate.rank - 1) * 2
            thumb_path = resolve_thumbnail_for_xlsx(candidate.thumbnail, output_dir)
            if not thumb_path.exists():
                continue
            try:
                image = XlsxImage(str(thumb_path))
                image.width = 92
                image.height = 132
                sheet.add_image(image, f"{get_column_letter(image_column)}{row_index}")
            except Exception:
                continue
        for column in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_index, column=column)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in range(candidate_text_start, candidate_text_start + max_candidates * 2, 2):
            sheet.cell(row=row_index, column=column).alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        1: 8,
        2: 20,
        3: 34,
        4: 18,
        5: 18,
        6: 16,
        7: 12,
    }
    for index in range(1, max_candidates + 1):
        widths[8 + (index - 1) * 2] = 16
        widths[9 + (index - 1) * 2] = 32
    tail_start = 8 + max_candidates * 2
    widths[tail_start] = 14
    widths[tail_start + 1] = 14
    widths[tail_start + 2] = 18
    widths[tail_start + 3] = 32
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def resolve_thumbnail_for_xlsx(thumbnail: str, output_dir: Path) -> Path:
    path = Path(thumbnail)
    if path.is_absolute():
        return path
    return output_dir / thumbnail.replace("/", "\\")


def write_feedback_template(csv_path: Path, json_path: Path, rows: list[MatchBoardRow]) -> None:
    fields = [
        "镜号",
        "字幕台词",
        "AI画面需求",
        "目标分类",
        "目标关键词",
        "当前状态",
        "确认结果",
        "选中素材序号",
        "替换关键词",
        "备注",
    ]
    records = []
    for row in rows:
        record = {
            "镜号": f"{row.index:03d}",
            "字幕台词": row.script_text,
            "AI画面需求": row.visual_need,
            "目标分类": row.target_category,
            "目标关键词": row.target_keyword,
            "当前状态": row.status,
            "确认结果": "",
            "选中素材序号": "",
            "替换关键词": "",
            "备注": "",
        }
        records.append(record)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(records)
    json_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def write_match_board_html(path: Path, title: str, rows: list[MatchBoardRow], output_dir: Path, audio_path: str = "") -> None:
    cards = []
    for row in rows:
        candidates = "\n".join(render_candidate_card(candidate) for candidate in row.candidates) or '<div class="empty">没有直接候选，先补素材或改关键词</div>'
        time_text = ""
        if row.start is not None or row.end is not None:
            time_text = f"{seconds_to_timestamp(row.start or 0)} - {seconds_to_timestamp(row.end or 0)}"
        cards.append(
            f"""
            <section class="beat" data-row="{row.index:03d}">
              <div class="script">
                <div class="index">{row.index:03d}</div>
                <div class="time">{html.escape(time_text)}</div>
                <h2>{html.escape(row.script_text)}</h2>
                <p><b>AI画面需求</b> {html.escape(row.visual_need)}</p>
                <p><b>目标</b> {html.escape(row.target_category)} / {html.escape(row.target_keyword or "未细分")}</p>
                <span class="status">{html.escape(row.status)}</span>
              </div>
              <div class="candidates">{candidates}</div>
              <textarea data-row="{row.index:03d}" placeholder="这里写你的备注：通过 / 替换 / 废料 / 重配原因"></textarea>
            </section>
            """
        )
    audio_block = f'<audio controls src="{html.escape(relative_or_absolute(Path(audio_path), output_dir))}"></audio>' if audio_path else ""
    document = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{html.escape(title)} 智能镜头匹配审片板</title>
  <style>
    :root {{ --bg:#e8f0f6; --panel:#eef4f8; --ink:#1c2938; --muted:#607083; --accent:#307eff; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:"Microsoft YaHei UI","Microsoft YaHei",system-ui,sans-serif; background:linear-gradient(145deg,#f4f8fb,#dfe9f1); color:var(--ink); }}
    header {{ position:sticky; top:0; z-index:2; padding:18px 24px; background:rgba(238,244,248,.86); backdrop-filter:blur(10px); box-shadow:0 1px 0 rgba(255,255,255,.86),0 10px 24px rgba(98,120,140,.12); }}
    h1 {{ margin:0 0 6px; font-size:24px; }}
    header p {{ margin:0; color:var(--muted); }}
    audio {{ width:min(720px,100%); margin-top:12px; }}
    .toolbar {{ display:flex; flex-wrap:wrap; gap:10px; margin-top:12px; }}
    button {{ border:0; border-radius:999px; padding:9px 14px; background:#307eff; color:#fff; font:inherit; cursor:pointer; box-shadow:4px 6px 14px rgba(48,126,255,.22),-3px -3px 10px rgba(255,255,255,.62); }}
    button.secondary {{ background:#edf4f8; color:#1c2938; border:1px solid rgba(255,255,255,.78); }}
    main {{ padding:18px; display:grid; gap:14px; }}
    .beat {{ display:grid; grid-template-columns:minmax(260px,360px) 1fr minmax(220px,300px); gap:14px; align-items:stretch; background:#e3edf4; border:1px solid rgba(255,255,255,.78); border-radius:22px; padding:14px; box-shadow:8px 12px 24px rgba(112,130,150,.18),-6px -6px 16px rgba(255,255,255,.75); }}
    .script, .candidate, textarea {{ border-radius:18px; border:1px solid rgba(255,255,255,.78); background:var(--panel); box-shadow:4px 6px 14px rgba(112,130,150,.14),-4px -4px 12px rgba(255,255,255,.70); }}
    .script {{ padding:14px; }}
    .index {{ font-weight:800; color:var(--accent); font-size:20px; }}
    .time {{ color:var(--muted); font-size:12px; min-height:18px; }}
    h2 {{ margin:10px 0; font-size:17px; line-height:1.55; }}
    .script p {{ margin:8px 0; color:var(--muted); line-height:1.45; }}
    .status {{ display:inline-flex; margin-top:8px; padding:5px 10px; border-radius:999px; background:#dbe8f8; color:#185bd7; font-size:12px; }}
    .candidates {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(150px,1fr)); gap:10px; }}
    .candidate {{ padding:8px; min-width:0; }}
    .candidate.direct {{ outline:2px solid rgba(48,126,255,.46); }}
    .candidate img {{ width:100%; aspect-ratio:9/13; object-fit:cover; display:block; border-radius:13px; background:#d5dee8; }}
    .candidate b {{ display:block; margin-top:7px; font-size:12px; line-height:1.35; word-break:break-word; }}
    .candidate small {{ display:block; margin-top:5px; color:var(--muted); line-height:1.35; }}
    textarea {{ width:100%; min-height:100%; padding:12px; resize:vertical; color:var(--ink); font-family:inherit; line-height:1.5; }}
    .empty {{ padding:20px; color:var(--muted); border-radius:18px; background:var(--panel); }}
    @media (max-width: 1100px) {{ .beat {{ grid-template-columns:1fr; }} textarea {{ min-height:120px; }} }}
  </style>
</head>
<body>
  <header>
    <h1>{html.escape(title)} 智能镜头匹配审片板</h1>
    <p>左边是台词，中间是候选素材截图，右边写备注。先审匹配，再进粗剪。</p>
    {audio_block}
    <div class="toolbar">
      <button onclick="exportFeedback()">导出审片反馈 JSON</button>
      <button class="secondary" onclick="clearFeedback()">清空本页备注</button>
    </div>
  </header>
  <main>
    {''.join(cards)}
  </main>
  <script>
    const boardKey = "match-board-feedback:" + location.pathname;
    const saved = JSON.parse(localStorage.getItem(boardKey) || "{{}}");
    document.querySelectorAll("textarea[data-row]").forEach(area => {{
      const row = area.dataset.row;
      area.value = saved[row]?.note || "";
      area.addEventListener("input", () => {{
        const current = JSON.parse(localStorage.getItem(boardKey) || "{{}}");
        current[row] = {{ row, note: area.value, updated_at: new Date().toISOString() }};
        localStorage.setItem(boardKey, JSON.stringify(current, null, 2));
      }});
    }});
    function exportFeedback(){{
      const payload = {{
        title: document.querySelector("h1")?.textContent || "",
        exported_at: new Date().toISOString(),
        rows: Object.values(JSON.parse(localStorage.getItem(boardKey) || "{{}}"))
      }};
      const blob = new Blob([JSON.stringify(payload, null, 2)], {{type:"application/json"}});
      const url = URL.createObjectURL(blob);
      const link = document.createElement("a");
      link.href = url;
      link.download = "智能镜头匹配审片反馈.json";
      link.click();
      URL.revokeObjectURL(url);
    }}
    function clearFeedback(){{
      if(confirm("清空当前审片板的本地备注？")) {{
        localStorage.removeItem(boardKey);
        document.querySelectorAll("textarea[data-row]").forEach(area => area.value = "");
      }}
    }}
  </script>
</body>
</html>"""
    path.write_text(document, encoding="utf-8")


def render_candidate_card(candidate: MatchCandidate) -> str:
    class_name = "candidate direct" if candidate.direct else "candidate"
    image = html.escape(candidate.thumbnail)
    return f"""
    <article class="{class_name}">
      <img src="{image}" alt="素材{candidate.rank}" />
      <b>{candidate.rank}. {html.escape(Path(candidate.clip_path).name)}</b>
      <small>{html.escape(candidate.category)} / {html.escape(candidate.keyword)}</small>
      <small>{html.escape(candidate.reason)}</small>
    </article>
    """


def row_to_json(row: MatchBoardRow, output_dir: Path) -> dict[str, object]:
    data = asdict(row)
    data["candidates"] = [asdict(candidate) for candidate in row.candidates]
    return data


def relative_or_absolute(path: Path, root: Path) -> str:
    if not path:
        return ""
    try:
        return str(path.resolve().relative_to(root.resolve())).replace("\\", "/")
    except Exception:
        return str(path)
